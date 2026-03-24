import io
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from babel.numbers import format_currency

from services.storage import load_csv, save_csv, get_path
from services.client_config import get_active_config
from process_functions import monetizacion_v0 as mon


# Nombres de los archivos persistentes
HISTORIAL_FILE = "airbnb_reservations_paid_historico.csv"
HISTORIAL_PAYOUT_FILE = "airbnb_payouts_historico.csv"
HISTORIAL_MONETIZACION_FILE = "monetizaciones.csv"


def run():
    config = get_active_config()
    client_id = config.client_id

    st.title("Monetizaciones")

    CUENTAS = config.monetizacion.cuentas

    # ----------------- SIDEBAR: ACTUALIZAR HISTORIAL AIRBNB -----------------
    with st.sidebar:
        st.subheader("Actualizar historial Airbnb")

        airbnb_file = st.file_uploader(
            "Sube uno o más archivos CSV de lotes",
            type=["csv"],
            accept_multiple_files=False,
        )

        if airbnb_file:
            cuenta = st.selectbox("Selecciona la cuenta", CUENTAS)

            if st.button("Actualizar historial"):
                clean_df = mon.fix_raw_data(airbnb_file, cuenta)

                updated_historical = mon.update_airbnb_historical(
                    clean_df,
                    get_path(HISTORIAL_FILE, client_id)
                )
                save_csv(HISTORIAL_FILE, updated_historical, client_id)

                payouts = mon.update_payout(
                    updated_historical,
                    get_path(HISTORIAL_PAYOUT_FILE, client_id)
                )
                save_csv(HISTORIAL_PAYOUT_FILE, payouts, client_id)

                st.success("Historial de payouts actualizado.")
        else:
            st.warning("NO SE HA CARGADO ARCHIVO DE AIRBNB")

    # ----------------- CARGAR HISTORIALES -----------------
    airbnb_historical = load_csv(HISTORIAL_FILE, client_id)
    payouts = load_csv(HISTORIAL_PAYOUT_FILE, client_id)

    # ----------------- AÑADIR PAYOUT MANUAL -----------------
    st.subheader("Agregar payout manual")

    # Auto-generar siguiente código Stripe
    stripe_nums = (
        payouts["Código de referencia"]
        .astype(str)
        .str.extract(r"^Stripe-(\d+)$", expand=False)
        .dropna()
    )
    stripe_default = config.monetizacion.stripe_default_start
    last_num = max(stripe_nums.astype(int).max(), stripe_default) if not stripe_nums.empty else stripe_default
    next_stripe_code = f"Stripe-{last_num + 1}"

    with st.expander("Agregar payout manual"):
        with st.form("form_nuevo_payout"):
            col_tipo, col_cuenta = st.columns(2)
            with col_tipo:
                tipo_payout = st.selectbox(
                    "Tipo de payout",
                    ["Stripe", "Otro"],
                    help="Stripe auto-genera el código. Otro permite código libre.",
                )
            with col_cuenta:
                fuente_payout = st.selectbox("Cuenta / Fuente", CUENTAS)

            if tipo_payout == "Stripe":
                st.markdown(
                    f"**Código de referencia automático:** `{next_stripe_code}`"
                )
                codigo_referencia = next_stripe_code
            else:
                codigo_referencia = st.text_input(
                    "Código de referencia",
                    placeholder="Ej: Boulevard, Transferencia-01, etc.",
                )

            fecha_nueva = st.date_input("Fecha de payout")
            col_usd, col_cop = st.columns(2)
            with col_usd:
                total_pagado_nuevo = st.number_input("Total pagado (USD)", min_value=0.0, format="%.2f")
            with col_cop:
                monto_nuevo = st.number_input("Monto (COP)", min_value=0.0, format="%.0f")
            numero_reservas_nueva = st.number_input("Número de reservas", min_value=0, step=1)

            submitted = st.form_submit_button("Guardar payout manual")
            if submitted:
                if not codigo_referencia or not codigo_referencia.strip():
                    st.error("El código de referencia no puede estar vacío.")
                elif total_pagado_nuevo <= 0:
                    st.error("El total pagado (USD) debe ser mayor a 0.")
                else:
                    new_row = {
                        "id_accival": "pre_accival",
                        "Código de referencia": codigo_referencia.strip(),
                        "Monto": monto_nuevo,
                        "Total pagado": total_pagado_nuevo,
                        "Fecha": fecha_nueva.strftime("%Y-%m-%d"),
                        "Cantidad reservas": numero_reservas_nueva,
                        "fuente": fuente_payout,
                    }

                    payouts = pd.concat([payouts, pd.DataFrame([new_row])], ignore_index=True)
                    save_csv(HISTORIAL_PAYOUT_FILE, payouts, client_id)

                    st.success(f"Payout manual agregado: **{codigo_referencia}** ({fuente_payout})")

    # ----------------- EDITAR id_accival -----------------
    st.subheader("Transacciones pendientes de monetización")
    st.caption("Actualizar ID Accival cuando se monetiza la transacción")

    payouts['id_accival'] = payouts['id_accival'].astype(str)

    df_editables = payouts[
        (payouts['id_accival'] == '') |
        (payouts['id_accival'].str.startswith('pre_accival'))
    ].copy()

    df_editables = df_editables[
        ['id_accival'] + [c for c in df_editables.columns if c != 'id_accival']
    ]

    df_editado = st.data_editor(
        df_editables,
        column_config={"id_accival": st.column_config.TextColumn("ID Accival")},
        disabled=[c for c in df_editables.columns if c != 'id_accival'],
        num_rows="fixed",
    )

    if st.button("Guardar cambios"):
        for idx, row in df_editado.iterrows():
            payouts.loc[payouts['Código de referencia'] == row['Código de referencia'], 'id_accival'] = row['id_accival']

        save_csv(HISTORIAL_PAYOUT_FILE, payouts, client_id)
        st.success("Cambios guardados correctamente.")

    # ----------------- ACTUALIZAR MONETIZACIÓN -----------------
    monet = mon.update_monetizacion(payouts, get_path(HISTORIAL_MONETIZACION_FILE, client_id))

    # Sincronizar
    monet["id_accival"] = monet["id_accival"].astype(str)
    payouts["id_accival"] = payouts["id_accival"].astype(str)

    pre_refs_payouts = payouts[payouts["id_accival"].str.startswith("pre_accival")]["id_accival"]

    if pre_refs_payouts.empty:
        monet = monet[~monet["id_accival"].str.startswith("pre_accival")]
    else:
        monet = monet[~(
            monet["id_accival"].str.startswith("pre_accival") &
            ~monet["id_accival"].isin(pre_refs_payouts)
        )]

    if "fecha" in monet.columns:
        monet["fecha"] = pd.to_datetime(monet["fecha"], errors="coerce")
        monet = monet.sort_values("fecha").drop_duplicates("id_accival", keep="last")

    save_csv(HISTORIAL_MONETIZACION_FILE, monet, client_id)

    # ----------------- LISTA DE MONETIZACIONES -----------------
    st.subheader("Registro histórico de monetizaciones")
    monet = monet.sort_values("fecha", ascending=False).reset_index(drop=True)
    st.dataframe(monet)

    # =====================================================================
    #  EXPORT PRE_ACCIVAL
    # =====================================================================
    df_pre_payout = payouts[payouts['id_accival'].str.startswith('pre_accival')][
        ['id_accival', 'Fecha', 'Código de referencia', 'Total pagado']
    ].copy()

    if not df_pre_payout.empty:
        from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

        buffer = io.BytesIO()
        df_pre_payout['Total pagado'] = pd.to_numeric(df_pre_payout['Total pagado'], errors='coerce')

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_pre_payout.to_excel(writer, index=False, sheet_name='pre_accival')
            ws = writer.sheets['pre_accival']

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            total_font = Font(bold=True, size=12)

            # Formato headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Bordes en todas las celdas de datos
            num_rows = len(df_pre_payout)
            num_cols = len(df_pre_payout.columns)
            for row in ws.iter_rows(min_row=2, max_row=num_rows + 1, max_col=num_cols):
                for cell in row:
                    cell.border = thin_border

            # Formato USD para columna "Total pagado" (columna D = 4)
            for row in ws.iter_rows(min_row=2, max_row=num_rows + 1, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = '$#,##0.00'

            # Fila de TOTAL
            total_row = num_rows + 2
            ws.cell(row=total_row, column=3, value='TOTAL').font = total_font
            ws.cell(row=total_row, column=3).border = thin_border
            ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')

            total_value = df_pre_payout['Total pagado'].sum()
            total_cell = ws.cell(row=total_row, column=4, value=total_value)
            total_cell.font = total_font
            total_cell.border = thin_border
            total_cell.number_format = '$#,##0.00'

            # Ajustar ancho de columnas
            col_widths = {'A': 16, 'B': 14, 'C': 24, 'D': 16}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        buffer.seek(0)

        st.download_button(
            label="Descargar Excel de pre_accival",
            data=buffer,
            file_name="pre_accival_transacciones.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("No hay filas con 'pre_accival' para exportar.")

    # =====================================================================
    #  ESTADÍSTICAS Y MÉTRICAS
    # =====================================================================
    st.header("📊 Estadísticas")

    # --- Preparar datos base ---
    df_pre = monet[monet['id_accival'].str.startswith('pre_accival')].copy()
    df_mon = monet[monet['id_accival'].str.startswith('accival')].copy()
    df_mon['fecha'] = pd.to_datetime(df_mon['fecha'], errors='coerce')

    # Convertir columnas numéricas
    for col in ['monto_usd', 'monto_cop', 'costo', 'trm', 'trm_dia']:
        if col in df_mon.columns:
            df_mon[col] = pd.to_numeric(df_mon[col], errors='coerce')
        if col in df_pre.columns:
            df_pre[col] = pd.to_numeric(df_pre[col], errors='coerce')

    # --- Filtro de fechas ---
    if not df_mon.empty and df_mon['fecha'].notna().any():
        fecha_min = df_mon['fecha'].min().date()
        fecha_max = df_mon['fecha'].max().date()

        col_f1, col_f2 = st.columns(2)
        with col_f1:
            fecha_inicio = st.date_input("Desde", value=fecha_min, min_value=fecha_min, max_value=fecha_max)
        with col_f2:
            fecha_fin = st.date_input("Hasta", value=fecha_max, min_value=fecha_min, max_value=fecha_max)

        mask_fecha = (df_mon['fecha'].dt.date >= fecha_inicio) & (df_mon['fecha'].dt.date <= fecha_fin)
        df_mon_filtered = df_mon[mask_fecha].copy()
    else:
        df_mon_filtered = df_mon.copy()
        fecha_inicio = None
        fecha_fin = None

    # --- KPIs en 2 filas ---
    st.subheader("Indicadores clave")

    # Calcular deltas (mes actual vs mes anterior)
    if not df_mon_filtered.empty:
        df_mon_filtered['mes'] = df_mon_filtered['fecha'].dt.to_period('M')
        meses_ordenados = sorted(df_mon_filtered['mes'].dropna().unique())

        if len(meses_ordenados) >= 2:
            mes_actual = meses_ordenados[-1]
            mes_anterior = meses_ordenados[-2]
            datos_actual = df_mon_filtered[df_mon_filtered['mes'] == mes_actual]
            datos_anterior = df_mon_filtered[df_mon_filtered['mes'] == mes_anterior]

            delta_usd = datos_actual['monto_usd'].sum() - datos_anterior['monto_usd'].sum()
            delta_cop = datos_actual['monto_cop'].sum() - datos_anterior['monto_cop'].sum()
            delta_costo = datos_actual['costo'].sum() - datos_anterior['costo'].sum()
        else:
            delta_usd = delta_cop = delta_costo = None

        total_usd = df_mon_filtered['monto_usd'].sum()
        total_cop = df_mon_filtered['monto_cop'].sum()
        total_costo = df_mon_filtered['costo'].sum()
    else:
        total_usd = total_cop = total_costo = 0
        delta_usd = delta_cop = delta_costo = None

    # Fila 1: Monetizado USD, Monetizado COP, Costo COP
    kpi_r1_c1, kpi_r1_c2, kpi_r1_c3 = st.columns(3)
    with kpi_r1_c1:
        st.metric(
            "MONETIZADO USD",
            format_currency(total_usd, "USD", locale='en_US'),
            delta=format_currency(delta_usd, "USD", locale='en_US') if delta_usd is not None else None,
        )
    with kpi_r1_c2:
        st.metric(
            "MONETIZADO COP",
            format_currency(total_cop, "COP", locale='es_CO'),
            delta=format_currency(delta_cop, "COP", locale='es_CO') if delta_cop is not None else None,
        )
    with kpi_r1_c3:
        st.metric(
            "COSTO COP",
            format_currency(total_costo, "COP", locale='es_CO'),
            delta=format_currency(delta_costo, "COP", locale='es_CO') if delta_costo is not None else None,
            delta_color="inverse",
        )

    # Fila 2: Por monetizar, TRM promedio (últimas 3), Ingreso potencial
    por_monetizar_usd = df_pre['monto_usd'].sum() if not df_pre.empty else 0

    # TRM promedio de las últimas 3 monetizaciones (no filtradas por fecha)
    df_mon_sorted = df_mon.dropna(subset=['fecha']).sort_values('fecha', ascending=False)
    trm_ultimas_3 = df_mon_sorted.head(3)['trm'].mean() if len(df_mon_sorted) >= 1 else 0

    ingreso_potencial = por_monetizar_usd * trm_ultimas_3

    kpi_r2_c1, kpi_r2_c2, kpi_r2_c3 = st.columns(3)
    with kpi_r2_c1:
        st.metric("POR MONETIZAR USD", format_currency(por_monetizar_usd, "USD", locale='en_US'))
    with kpi_r2_c2:
        st.metric("TRM PROMEDIO (últ. 3)", format_currency(trm_ultimas_3, "COP", locale='es_CO'))
    with kpi_r2_c3:
        st.metric("INGRESO POTENCIAL COP", format_currency(ingreso_potencial, "COP", locale='es_CO'))

    # =====================================================================
    #  TABLA MENSUAL — Desglose Airbnb (por cuenta) vs Stripe
    # =====================================================================
    st.subheader("Histórico Mensual — Desglose por fuente")

    # Clasificar payouts: Airbnb (G-*) vs Stripe (el resto)
    payouts_con_fecha = payouts.copy()
    payouts_con_fecha['Fecha'] = pd.to_datetime(payouts_con_fecha['Fecha'], errors='coerce')
    payouts_con_fecha['Total pagado'] = pd.to_numeric(payouts_con_fecha['Total pagado'], errors='coerce')
    payouts_con_fecha = payouts_con_fecha.dropna(subset=['Fecha'])
    payouts_con_fecha['mes'] = payouts_con_fecha['Fecha'].dt.to_period('M')

    # Aplicar filtro de fechas también a payouts
    if fecha_inicio and fecha_fin:
        mask_payout_fecha = (
            (payouts_con_fecha['Fecha'].dt.date >= fecha_inicio) &
            (payouts_con_fecha['Fecha'].dt.date <= fecha_fin)
        )
        payouts_filtrados = payouts_con_fecha[mask_payout_fecha].copy()
    else:
        payouts_filtrados = payouts_con_fecha.copy()

    cod_ref = payouts_filtrados['Código de referencia'].astype(str)
    payouts_filtrados['es_airbnb'] = cod_ref.str.strip().str.startswith('G-')
    payouts_filtrados['tipo'] = payouts_filtrados['es_airbnb'].map({True: 'Airbnb', False: 'Stripe'})

    # Detectar cuentas Airbnb dinámicamente
    airbnb_payouts = payouts_filtrados[payouts_filtrados['es_airbnb']].copy()
    airbnb_payouts['fuente'] = airbnb_payouts['fuente'].astype(str).str.strip()
    cuentas_airbnb = sorted(airbnb_payouts['fuente'].unique()) if not airbnb_payouts.empty else []

    # Pivot Airbnb por cuenta y mes
    if not airbnb_payouts.empty:
        airbnb_por_cuenta = airbnb_payouts.pivot_table(
            index='mes',
            columns='fuente',
            values='Total pagado',
            aggfunc='sum',
            fill_value=0,
        )
        airbnb_por_cuenta.columns = [f"Airbnb {c}" for c in airbnb_por_cuenta.columns]
        airbnb_por_cuenta['Airbnb Total USD'] = airbnb_por_cuenta.sum(axis=1)
    else:
        airbnb_por_cuenta = pd.DataFrame()

    # Stripe por mes
    stripe_payouts = payouts_filtrados[~payouts_filtrados['es_airbnb']].copy()
    if not stripe_payouts.empty:
        stripe_mensual = stripe_payouts.groupby('mes')['Total pagado'].sum().rename('Stripe USD')
    else:
        stripe_mensual = pd.Series(dtype=float, name='Stripe USD')

    # Tabla de monetizaciones mensual (para TRM y transacciones)
    if not df_mon_filtered.empty:
        mon_mensual = df_mon_filtered.groupby('mes').agg(
            transacciones=('id_accival', 'count'),
            trm_promedio=('trm', 'mean'),
            monto_usd=('monto_usd', 'sum'),
            monto_cop=('monto_cop', 'sum'),
            costo=('costo', 'sum'),
        )
        mon_mensual['trm_promedio'] = mon_mensual['trm_promedio'].round(0)
    else:
        mon_mensual = pd.DataFrame()

    # Combinar todo
    tablas = [df for df in [airbnb_por_cuenta, stripe_mensual.to_frame(), mon_mensual] if not df.empty]
    if tablas:
        tabla_final = tablas[0]
        for t in tablas[1:]:
            tabla_final = tabla_final.join(t, how='outer')
        tabla_final = tabla_final.fillna(0).sort_index()

        # Calcular Total USD si hay Airbnb y Stripe
        if 'Airbnb Total USD' in tabla_final.columns and 'Stripe USD' in tabla_final.columns:
            tabla_final['Total USD'] = tabla_final['Airbnb Total USD'] + tabla_final['Stripe USD']
        elif 'Airbnb Total USD' in tabla_final.columns:
            tabla_final['Total USD'] = tabla_final['Airbnb Total USD']
        elif 'Stripe USD' in tabla_final.columns:
            tabla_final['Total USD'] = tabla_final['Stripe USD']

        # Reordenar columnas
        col_order = []
        # Cuentas Airbnb individuales primero
        for c in sorted([c for c in tabla_final.columns if c.startswith('Airbnb ') and c != 'Airbnb Total USD']):
            col_order.append(c)
        if 'Airbnb Total USD' in tabla_final.columns:
            col_order.append('Airbnb Total USD')
        if 'Stripe USD' in tabla_final.columns:
            col_order.append('Stripe USD')
        if 'Total USD' in tabla_final.columns:
            col_order.append('Total USD')
        if 'transacciones' in tabla_final.columns:
            col_order.append('transacciones')
        if 'trm_promedio' in tabla_final.columns:
            col_order.append('trm_promedio')
        if 'monto_cop' in tabla_final.columns:
            col_order.append('monto_cop')
        if 'costo' in tabla_final.columns:
            col_order.append('costo')
        # Excluir monto_usd duplicado (ya tenemos Total USD del desglose)
        col_order = [c for c in col_order if c in tabla_final.columns]
        tabla_display = tabla_final[col_order].copy()

        # Formato para display
        tabla_display.index = tabla_display.index.astype(str)
        tabla_display.index.name = 'Mes'

        # Formatear columnas monetarias
        cols_usd = [c for c in tabla_display.columns if 'USD' in c or c == 'monto_usd']
        cols_cop = [c for c in tabla_display.columns if c in ['monto_cop', 'costo']]

        tabla_fmt = tabla_display.copy()
        for c in cols_usd:
            tabla_fmt[c] = tabla_fmt[c].apply(lambda x: f"${x:,.0f}")
        for c in cols_cop:
            tabla_fmt[c] = tabla_fmt[c].apply(lambda x: f"${x:,.0f}")
        if 'trm_promedio' in tabla_fmt.columns:
            tabla_fmt['trm_promedio'] = tabla_fmt['trm_promedio'].apply(lambda x: f"${x:,.0f}")
        if 'transacciones' in tabla_fmt.columns:
            tabla_fmt['transacciones'] = tabla_fmt['transacciones'].astype(int)

        st.dataframe(tabla_fmt, use_container_width=True)
    else:
        st.info("No hay datos suficientes para la tabla mensual.")

    # =====================================================================
    #  GRÁFICOS CON PLOTLY
    # =====================================================================
    st.subheader("📈 Gráficas de comportamiento")

    # --- 1. Barras apiladas mensuales por fuente ---
    if not payouts_filtrados.empty:
        # Preparar datos para barras: Airbnb por cuenta + Stripe
        payouts_chart = payouts_filtrados.copy()
        payouts_chart['fuente_display'] = payouts_chart.apply(
            lambda r: f"Airbnb {r['fuente']}" if r['es_airbnb'] else 'Stripe', axis=1
        )
        payouts_chart['mes_str'] = payouts_chart['mes'].astype(str)

        chart_data = payouts_chart.groupby(['mes_str', 'fuente_display'])['Total pagado'].sum().reset_index()

        fig_barras = px.bar(
            chart_data,
            x='mes_str',
            y='Total pagado',
            color='fuente_display',
            title='Ingresos mensuales por fuente (USD)',
            labels={'mes_str': 'Mes', 'Total pagado': 'USD', 'fuente_display': 'Fuente'},
            barmode='stack',
        )
        fig_barras.update_layout(
            xaxis_title='Mes',
            yaxis_title='USD',
            yaxis_tickformat='$,.0f',
            legend_title='Fuente',
            hovermode='x unified',
        )
        st.plotly_chart(fig_barras, use_container_width=True)

    # --- 2. Línea TRM y TRM del día ---
    if not df_mon_filtered.empty:
        df_trm = df_mon_filtered.dropna(subset=['fecha']).sort_values('fecha')

        if not df_trm.empty:
            fig_trm = go.Figure()
            fig_trm.add_trace(go.Scatter(
                x=df_trm['fecha'],
                y=df_trm['trm'],
                mode='lines+markers',
                name='TRM Monetización',
                line=dict(color='#636EFA', width=2),
                hovertemplate='%{x|%d %b %Y}<br>TRM: $%{y:,.0f}<extra></extra>',
            ))

            trm_dia = pd.to_numeric(df_trm['trm_dia'], errors='coerce')
            if trm_dia.notna().any():
                fig_trm.add_trace(go.Scatter(
                    x=df_trm['fecha'],
                    y=trm_dia,
                    mode='lines+markers',
                    name='TRM del Día',
                    line=dict(color='#EF553B', width=2, dash='dash'),
                    hovertemplate='%{x|%d %b %Y}<br>TRM Día: $%{y:,.0f}<extra></extra>',
                ))

            fig_trm.update_layout(
                title='Evolución TRM',
                xaxis_title='Fecha',
                yaxis_title='COP por USD',
                yaxis_tickformat='$,.0f',
                hovermode='x unified',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            )
            st.plotly_chart(fig_trm, use_container_width=True)

    # --- 3. Monto USD acumulado ---
    if not df_mon_filtered.empty:
        df_acum = df_mon_filtered.dropna(subset=['fecha']).sort_values('fecha').copy()

        if not df_acum.empty:
            df_acum['usd_acumulado'] = df_acum['monto_usd'].cumsum()

            fig_acum = go.Figure()
            fig_acum.add_trace(go.Scatter(
                x=df_acum['fecha'],
                y=df_acum['usd_acumulado'],
                mode='lines',
                fill='tozeroy',
                name='USD Acumulado',
                line=dict(color='#00CC96', width=2),
                hovertemplate='%{x|%d %b %Y}<br>Acumulado: $%{y:,.0f}<extra></extra>',
            ))

            fig_acum.update_layout(
                title='Monetización USD acumulada',
                xaxis_title='Fecha',
                yaxis_title='USD',
                yaxis_tickformat='$,.0f',
                hovermode='x unified',
            )
            st.plotly_chart(fig_acum, use_container_width=True)

    # =====================================================================
    #  PAYOUTS
    # =====================================================================
    st.subheader("Payouts")
    st.caption("Listado de todos los payouts enviados con su id de la monetización de Accival")
    st.dataframe(payouts, use_container_width=True)

    # =====================================================================
    #  MODO EDITOR — Corrección de datos
    # =====================================================================
    st.header("Modo Editor — Corrección de datos")
    st.caption("Edita cualquier campo en caso de error. Los cambios se guardan directamente en el volume.")

    tab_payouts, tab_monet, tab_reserv = st.tabs(["Payouts", "Monetizaciones", "Reservaciones"])

    # --- Tab 1: Editor de Payouts ---
    with tab_payouts:
        st.markdown("**Todas las columnas son editables.** Usa el checkbox para marcar filas a eliminar.")

        payouts_edit = payouts.copy()
        payouts_edit['Fecha'] = payouts_edit['Fecha'].astype(str).replace('NaT', '')
        payouts_edit.insert(0, '_eliminar', False)

        payouts_editado = st.data_editor(
            payouts_edit,
            column_config={
                "_eliminar": st.column_config.CheckboxColumn("Eliminar", default=False),
                "id_accival": st.column_config.TextColumn("ID Accival"),
                "Código de referencia": st.column_config.TextColumn("Código Ref."),
                "Monto": st.column_config.NumberColumn("Monto COP", format="%.0f"),
                "Total pagado": st.column_config.NumberColumn("Total USD", format="%.2f"),
                "Fecha": st.column_config.TextColumn("Fecha"),
                "Cantidad reservas": st.column_config.NumberColumn("# Reservas", format="%d"),
                "fuente": st.column_config.TextColumn("Fuente"),
            },
            num_rows="fixed",
            use_container_width=True,
            key="editor_payouts",
        )

        col_save_p, col_del_p = st.columns(2)
        with col_save_p:
            if st.button("Guardar cambios (Payouts)", key="btn_save_payouts"):
                df_guardado = payouts_editado.drop(columns=['_eliminar'])
                save_csv(HISTORIAL_PAYOUT_FILE, df_guardado, client_id)
                st.success("Payouts actualizados correctamente.")
                st.rerun()
        with col_del_p:
            filas_eliminar = payouts_editado[payouts_editado['_eliminar'] == True]
            if st.button(
                f"Eliminar {len(filas_eliminar)} fila(s) seleccionada(s)",
                key="btn_del_payouts",
                disabled=len(filas_eliminar) == 0,
            ):
                df_sin_eliminados = payouts_editado[payouts_editado['_eliminar'] != True].drop(columns=['_eliminar'])
                save_csv(HISTORIAL_PAYOUT_FILE, df_sin_eliminados, client_id)
                st.success(f"{len(filas_eliminar)} fila(s) eliminada(s) de payouts.")
                st.rerun()

    # --- Tab 2: Editor de Monetizaciones ---
    with tab_monet:
        st.markdown("**Edita trm, monto_cop, costo, fecha y más.** Usa el checkbox para marcar filas a eliminar.")

        monet_edit = monet.copy()
        monet_edit['fecha'] = monet_edit['fecha'].astype(str).replace('NaT', '')
        for _col in ['monto_usd', 'num_transacciones', 'trm', 'monto_cop', 'costo', 'trm_dia']:
            if _col in monet_edit.columns:
                monet_edit[_col] = pd.to_numeric(monet_edit[_col], errors='coerce')
        monet_edit.insert(0, '_eliminar', False)

        monet_editado = st.data_editor(
            monet_edit,
            column_config={
                "_eliminar": st.column_config.CheckboxColumn("Eliminar", default=False),
                "id_accival": st.column_config.TextColumn("ID Accival"),
                "monto_usd": st.column_config.NumberColumn("USD", format="%.2f"),
                "num_transacciones": st.column_config.NumberColumn("# Trans."),
                "trm": st.column_config.NumberColumn("TRM", format="%.0f"),
                "monto_cop": st.column_config.NumberColumn("COP", format="%.0f"),
                "costo": st.column_config.NumberColumn("Costo", format="%.0f"),
                "fecha": st.column_config.TextColumn("Fecha"),
                "trm_dia": st.column_config.NumberColumn("TRM Día", format="%.0f"),
            },
            num_rows="fixed",
            use_container_width=True,
            key="editor_monet",
        )

        col_save_m, col_del_m, col_recalc = st.columns(3)
        with col_save_m:
            if st.button("Guardar cambios (Monetizaciones)", key="btn_save_monet"):
                df_guardado = monet_editado.drop(columns=['_eliminar'])
                save_csv(HISTORIAL_MONETIZACION_FILE, df_guardado, client_id)
                st.success("Monetizaciones actualizadas correctamente.")
                st.rerun()
        with col_del_m:
            filas_eliminar_m = monet_editado[monet_editado['_eliminar'] == True]
            if st.button(
                f"Eliminar {len(filas_eliminar_m)} fila(s)",
                key="btn_del_monet",
                disabled=len(filas_eliminar_m) == 0,
            ):
                df_sin_eliminados = monet_editado[monet_editado['_eliminar'] != True].drop(columns=['_eliminar'])
                save_csv(HISTORIAL_MONETIZACION_FILE, df_sin_eliminados, client_id)
                st.success(f"{len(filas_eliminar_m)} fila(s) eliminada(s) de monetizaciones.")
                st.rerun()
        with col_recalc:
            if st.button("Recalcular COP y Costo", key="btn_recalc"):
                df_recalc = monet_editado.drop(columns=['_eliminar']).copy()
                df_recalc['trm'] = pd.to_numeric(df_recalc['trm'], errors='coerce')
                df_recalc['monto_usd'] = pd.to_numeric(df_recalc['monto_usd'], errors='coerce')
                df_recalc['trm_dia'] = pd.to_numeric(df_recalc['trm_dia'], errors='coerce')

                mask_recalc = df_recalc['trm'].notna() & df_recalc['monto_usd'].notna()
                df_recalc.loc[mask_recalc, 'monto_cop'] = (
                    df_recalc.loc[mask_recalc, 'monto_usd'] * df_recalc.loc[mask_recalc, 'trm']
                )

                mask_costo = mask_recalc & df_recalc['trm_dia'].notna()
                df_recalc.loc[mask_costo, 'costo'] = (
                    df_recalc.loc[mask_costo, 'monto_usd']
                    * (df_recalc.loc[mask_costo, 'trm'] - df_recalc.loc[mask_costo, 'trm_dia'])
                )

                save_csv(HISTORIAL_MONETIZACION_FILE, df_recalc, client_id)
                st.success("Recalculado: monto_cop = usd × trm, costo = usd × (trm − trm_dia)")
                st.rerun()

    # --- Tab 3: Editor de Reservaciones ---
    with tab_reserv:
        st.markdown("**Historial de reservaciones Airbnb.** Todas las columnas son editables.")

        reserv_edit = airbnb_historical.copy()
        if 'Fecha' in reserv_edit.columns:
            reserv_edit['Fecha'] = reserv_edit['Fecha'].astype(str).replace('NaT', '')
        reserv_edit.insert(0, '_eliminar', False)

        reserv_editado = st.data_editor(
            reserv_edit,
            column_config={
                "_eliminar": st.column_config.CheckboxColumn("Eliminar", default=False),
            },
            num_rows="fixed",
            use_container_width=True,
            key="editor_reserv",
        )

        col_save_r, col_del_r = st.columns(2)
        with col_save_r:
            if st.button("Guardar cambios (Reservaciones)", key="btn_save_reserv"):
                df_guardado = reserv_editado.drop(columns=['_eliminar'])
                save_csv(HISTORIAL_FILE, df_guardado, client_id)
                st.success("Reservaciones actualizadas correctamente.")
                st.rerun()
        with col_del_r:
            filas_eliminar_r = reserv_editado[reserv_editado['_eliminar'] == True]
            if st.button(
                f"Eliminar {len(filas_eliminar_r)} fila(s)",
                key="btn_del_reserv",
                disabled=len(filas_eliminar_r) == 0,
            ):
                df_sin_eliminados = reserv_editado[reserv_editado['_eliminar'] != True].drop(columns=['_eliminar'])
                save_csv(HISTORIAL_FILE, df_sin_eliminados, client_id)
                st.success(f"{len(filas_eliminar_r)} fila(s) eliminada(s) de reservaciones.")
                st.rerun()
